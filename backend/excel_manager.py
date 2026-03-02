import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "expenses.xlsx"

def save_expense(data):
    try:
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["Store", "Date", "Total", "Items"])
            wb.save(FILE_NAME)

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        items_str = ", ".join(data.get("items", []))

        ws.append([
            data.get("store"),
            data.get("date"),
            data.get("total"),
            items_str
        ])

        wb.save(FILE_NAME)

    except PermissionError:
        print("❌ Close expenses.xlsx before uploading new receipt")